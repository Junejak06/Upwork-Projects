import openpyxl
from openpyxl.styles import PatternFill

# File path for your Excel file
file_path = "/Users/kunaljuneja/Upwork/FinanceProject/nasdaq_colorcode.xlsx"

# Load the workbook and select the active sheet
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Conditional formatting colors
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

# Apply conditional formatting for 'Median Estimate' (assumed to be column F)
for row in range(2, sheet.max_row + 1):
    median_estimate_cell = f'F{row}'
    cell_value = str(sheet[median_estimate_cell].value)

    if cell_value.startswith('-'):
        sheet[median_estimate_cell].fill = red_fill
    elif cell_value.startswith('+'):
        sheet[median_estimate_cell].fill = green_fill

# Apply conditional formatting based on comparison between 'Intrinsic Value' and 'Previous Close'
# Assuming 'Intrinsic Value' is column I and 'Previous Close' is column J
for row in range(2, sheet.max_row + 1):
    intrinsic_value_cell = f'I{row}'
    previous_close_cell = f'J{row}'

    # Ensure both cells have numeric values
    if isinstance(sheet[intrinsic_value_cell].value, (int, float)) and isinstance(sheet[previous_close_cell].value, (int, float)):
        if sheet[intrinsic_value_cell].value > sheet[previous_close_cell].value:
            sheet[intrinsic_value_cell].fill = red_fill
        elif sheet[intrinsic_value_cell].value < sheet[previous_close_cell].value:
            sheet[intrinsic_value_cell].fill = green_fill

# Save the workbook
workbook.save(file_path)
